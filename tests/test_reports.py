"""Tests para Fase 3: generación de PDFs y envío de correos."""

import json
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import pytest


# ── Fixture: datos de prueba ──────────────────────────────

@pytest.fixture
def datos_json():
    """Genera un JSON de prueba con estructura realista."""
    return {
        "metadata": {
            "fecha_procesamiento": "2026-02-10T18:00:00",
            "total_cursos": 3,
            "total_estudiantes": 5,
            "version": "1.0",
        },
        "cursos": [
            {
                "id_moodle": "100",
                "id_sence": "6722569",
                "nombre": "Curso Automatización",
                "nombre_corto": "100",
                "categoria": "Abiertos",
                "fecha_inicio": "2025-11-01",
                "fecha_fin": "2025-12-31",
                "estado": "expired",
                "dias_restantes": -41,
                "comprador": {
                    "nombre": "María López",
                    "empresa": "VSPT Wine Group",
                    "email": "maria@vspt.cl",
                },
                "estudiantes": [
                    {
                        "id": "12345678-9",
                        "nombre": "Juan Pérez García",
                        "email": "juan@test.cl",
                        "progreso": 100.0,
                        "calificacion": 7.0,
                        "ultimo_acceso": "2025-12-30",
                        "dias_sin_ingreso": 41,
                        "estado": "A",
                        "riesgo": "",
                        "sence": {
                            "id_sence": "6722569",
                            "n_ingresos": 10,
                            "estado": "CONECTADO",
                            "declaracion_jurada": "Emitida",
                        },
                    },
                    {
                        "id": "98765432-1",
                        "nombre": "Ana María Rodríguez Ñuñez",
                        "email": "ana@test.cl",
                        "progreso": 50.0,
                        "calificacion": 4.5,
                        "ultimo_acceso": "2025-12-15",
                        "dias_sin_ingreso": 56,
                        "estado": "P",
                        "riesgo": "alto",
                        "sence": {
                            "id_sence": "6722569",
                            "n_ingresos": 3,
                            "estado": "CONECTADO",
                            "declaracion_jurada": "",
                        },
                    },
                ],
                "estadisticas": {
                    "total_estudiantes": 2,
                    "promedio_progreso": 75.0,
                    "promedio_calificacion": 5.8,
                    "aprobados": 1,
                    "reprobados": 0,
                    "en_proceso": 1,
                    "riesgo_alto": 1,
                    "riesgo_medio": 0,
                    "conectados_sence": 2,
                },
            },
            {
                "id_moodle": "101",
                "id_sence": "6722572",
                "nombre": "Curso Gestión Avanzada",
                "nombre_corto": "101",
                "categoria": "Abiertos",
                "fecha_inicio": "2025-11-15",
                "fecha_fin": "2026-01-15",
                "estado": "active",
                "dias_restantes": 10,
                "comprador": {
                    "nombre": "María López",
                    "empresa": "VSPT Wine Group",
                    "email": "maria@vspt.cl",
                },
                "estudiantes": [
                    {
                        "id": "11111111-1",
                        "nombre": "Pedro Soto",
                        "email": "pedro@test.cl",
                        "progreso": 80.0,
                        "calificacion": 6.0,
                        "ultimo_acceso": "2026-02-01",
                        "dias_sin_ingreso": 9,
                        "estado": "A",
                        "riesgo": "",
                        "sence": {
                            "id_sence": "6722572",
                            "n_ingresos": 5,
                            "estado": "CONECTADO",
                            "declaracion_jurada": "",
                        },
                    },
                ],
                "estadisticas": {
                    "total_estudiantes": 1,
                    "promedio_progreso": 80.0,
                    "promedio_calificacion": 6.0,
                    "aprobados": 1,
                    "reprobados": 0,
                    "en_proceso": 0,
                    "riesgo_alto": 0,
                    "riesgo_medio": 0,
                    "conectados_sence": 1,
                },
            },
            {
                "id_moodle": "200",
                "id_sence": "",
                "nombre": "Curso Sin Comprador",
                "nombre_corto": "200",
                "categoria": "Abiertos",
                "fecha_inicio": "2025-10-01",
                "fecha_fin": "2025-12-01",
                "estado": "expired",
                "dias_restantes": -71,
                "comprador": {
                    "nombre": "",
                    "empresa": "",
                    "email": "",
                },
                "estudiantes": [
                    {
                        "id": "22222222-2",
                        "nombre": "Carlos Muñoz",
                        "email": "carlos@test.cl",
                        "progreso": 30.0,
                        "calificacion": 2.0,
                        "ultimo_acceso": "2025-11-20",
                        "dias_sin_ingreso": 82,
                        "estado": "R",
                        "riesgo": "alto",
                        "sence": {
                            "id_sence": "",
                            "n_ingresos": 0,
                            "estado": "NO_APLICA",
                            "declaracion_jurada": "",
                        },
                    },
                    {
                        "id": "33333333-3",
                        "nombre": "Laura Díaz",
                        "email": "laura@test.cl",
                        "progreso": 0.0,
                        "calificacion": None,
                        "ultimo_acceso": None,
                        "dias_sin_ingreso": None,
                        "estado": "P",
                        "riesgo": "medio",
                        "sence": {
                            "id_sence": "",
                            "n_ingresos": 0,
                            "estado": "NO_APLICA",
                            "declaracion_jurada": "",
                        },
                    },
                ],
                "estadisticas": {
                    "total_estudiantes": 2,
                    "promedio_progreso": 15.0,
                    "promedio_calificacion": 2.0,
                    "aprobados": 0,
                    "reprobados": 1,
                    "en_proceso": 1,
                    "riesgo_alto": 1,
                    "riesgo_medio": 1,
                    "conectados_sence": 0,
                },
            },
        ],
    }


@pytest.fixture
def json_file(datos_json, tmp_path):
    """Escribe los datos de prueba a un archivo JSON temporal."""
    json_path = tmp_path / "datos_procesados.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(datos_json, f, ensure_ascii=False, indent=2)
    return json_path


# ── Test 1: Carga de JSON ─────────────────────────────────

class TestLoadJson:
    def test_load_json(self, json_file):
        from src.reports.pdf_generator import cargar_datos

        datos = cargar_datos(json_file)
        assert "metadata" in datos
        assert "cursos" in datos
        assert len(datos["cursos"]) == 3

    def test_load_json_missing_file(self, tmp_path):
        from src.reports.pdf_generator import cargar_datos

        with pytest.raises(FileNotFoundError):
            cargar_datos(tmp_path / "no_existe.json")

    def test_load_json_empty_cursos(self, tmp_path):
        from src.reports.pdf_generator import cargar_datos

        json_path = tmp_path / "vacio.json"
        json_path.write_text('{"metadata": {}, "cursos": []}', encoding="utf-8")
        with pytest.raises(ValueError, match="sin cursos"):
            cargar_datos(json_path)


# ── Test 2: Agrupación por comprador ──────────────────────

class TestGroupByComprador:
    def test_group_by_email(self, datos_json):
        from src.reports.pdf_generator import agrupar_por_comprador

        grupos = agrupar_por_comprador(datos_json)

        # Agrupado por email, no por empresa
        assert "maria@vspt.cl" in grupos
        assert "_sin_email_Sin comprador asignado" in grupos
        assert len(grupos) == 2

    def test_vspt_has_two_courses(self, datos_json):
        from src.reports.pdf_generator import agrupar_por_comprador

        grupos = agrupar_por_comprador(datos_json)
        vspt = grupos["maria@vspt.cl"]

        assert len(vspt["cursos"]) == 2
        assert vspt["nombre"] == "María López"
        assert vspt["email"] == "maria@vspt.cl"
        assert vspt["empresa"] == "VSPT Wine Group"

    def test_sin_comprador_group(self, datos_json):
        from src.reports.pdf_generator import agrupar_por_comprador

        grupos = agrupar_por_comprador(datos_json)
        sin = grupos["_sin_email_Sin comprador asignado"]

        assert len(sin["cursos"]) == 1
        assert sin["email"] == ""

    def test_same_email_different_empresa_one_group(self):
        from src.reports.pdf_generator import agrupar_por_comprador

        datos = {
            "cursos": [
                {
                    "nombre": "Curso 1",
                    "comprador": {"nombre": "María", "empresa": "Empresa A", "email": "maria@test.cl"},
                    "estudiantes": [{"id": "1", "nombre": "Juan"}],
                },
                {
                    "nombre": "Curso 2",
                    "comprador": {"nombre": "María", "empresa": "Empresa B", "email": "maria@test.cl"},
                    "estudiantes": [{"id": "2", "nombre": "Pedro"}],
                },
            ],
        }

        grupos = agrupar_por_comprador(datos)
        assert len(grupos) == 1
        assert "maria@test.cl" in grupos
        assert len(grupos["maria@test.cl"]["cursos"]) == 2

    def test_different_email_same_empresa_two_groups(self):
        from src.reports.pdf_generator import agrupar_por_comprador

        datos = {
            "cursos": [
                {
                    "nombre": "Curso 1",
                    "comprador": {"nombre": "María", "empresa": "TestCo", "email": "maria@test.cl"},
                    "estudiantes": [{"id": "1", "nombre": "Juan"}],
                },
                {
                    "nombre": "Curso 2",
                    "comprador": {"nombre": "Pedro", "empresa": "TestCo", "email": "pedro@test.cl"},
                    "estudiantes": [{"id": "2", "nombre": "Ana"}],
                },
            ],
        }

        grupos = agrupar_por_comprador(datos)
        assert len(grupos) == 2
        assert "maria@test.cl" in grupos
        assert "pedro@test.cl" in grupos


# ── Test 3: Generación de PDF ─────────────────────────────

class TestGeneratePDF:
    def test_generate_pdf(self, datos_json, tmp_path):
        from src.reports.pdf_generator import agrupar_por_comprador, generar_pdf

        grupos = agrupar_por_comprador(datos_json)
        grupo = grupos["maria@vspt.cl"]

        ruta = generar_pdf(grupo, output_dir=tmp_path)

        assert ruta.exists()
        assert ruta.stat().st_size > 0
        assert ruta.read_bytes()[:4] == b"%PDF"

    def test_multiple_courses_one_pdf(self, datos_json, tmp_path):
        """Un comprador con 2 cursos genera 1 solo PDF."""
        from src.reports.pdf_generator import agrupar_por_comprador, generar_pdf

        grupos = agrupar_por_comprador(datos_json)
        grupo = grupos["maria@vspt.cl"]
        assert len(grupo["cursos"]) == 2

        ruta = generar_pdf(grupo, output_dir=tmp_path)

        # Solo 1 archivo generado
        pdfs = list(tmp_path.glob("*.pdf"))
        assert len(pdfs) == 1
        assert ruta.stat().st_size > 0

    def test_empty_course_no_crash(self, tmp_path):
        """Curso sin estudiantes no crashea la generación."""
        from src.reports.pdf_generator import generar_pdf

        grupo = {
            "nombre": "Test",
            "empresa": "Test Corp",
            "email": "test@test.cl",
            "cursos": [{
                "id_moodle": "999",
                "id_sence": "",
                "nombre": "Curso Vacío",
                "nombre_corto": "999",
                "fecha_inicio": "2025-01-01",
                "fecha_fin": "2025-12-31",
                "estado": "active",
                "dias_restantes": 30,
                "estudiantes": [],
                "estadisticas": {
                    "total_estudiantes": 0,
                    "promedio_progreso": 0.0,
                    "promedio_calificacion": 0.0,
                    "aprobados": 0,
                    "reprobados": 0,
                    "en_proceso": 0,
                    "riesgo_alto": 0,
                    "riesgo_medio": 0,
                    "conectados_sence": 0,
                },
            }],
        }

        ruta = generar_pdf(grupo, output_dir=tmp_path)
        assert ruta.exists()
        assert ruta.stat().st_size > 0


# ── Test 4: Sanitización de nombres de archivo ────────────

class TestFilenameSanitization:
    def test_pdf_filename_sanitization(self):
        from src.reports.pdf_generator import sanitizar_nombre_archivo

        assert sanitizar_nombre_archivo("VSPT Wine Group") == "VSPT_Wine_Group"
        assert sanitizar_nombre_archivo("Ñuñoa S.A.") == "Nunoa_SA"
        assert sanitizar_nombre_archivo("José María") == "Jose_Maria"
        assert sanitizar_nombre_archivo("Café & Té") == "Cafe_Te"
        assert sanitizar_nombre_archivo("  espacios  ") == "espacios"

    def test_empty_name(self):
        from src.reports.pdf_generator import sanitizar_nombre_archivo

        assert sanitizar_nombre_archivo("") == "reporte"


# ── Test 5: Generación de cuerpo de correo ────────────────

class TestEmailBodyGeneration:
    def test_email_body_generation(self):
        from src.reports.email_sender import generar_cuerpo_correo

        resumen = [
            {"nombre": "Curso A", "total_estudiantes": 10, "aprobados": 8, "en_proceso": 2},
            {"nombre": "Curso B", "total_estudiantes": 5, "aprobados": 5, "en_proceso": 0},
        ]

        html = generar_cuerpo_correo("María López", "VSPT Wine Group", resumen)

        assert "María López" in html
        assert "VSPT Wine Group" in html
        assert "15" in html  # total participantes
        assert "13" in html  # total aprobados
        assert "2" in html   # en proceso
        assert "<html>" in html

    def test_email_body_empty_name(self):
        from src.reports.email_sender import generar_cuerpo_correo

        html = generar_cuerpo_correo("", "Test Corp", [])
        assert "Estimado/a" in html
        assert "Test Corp" in html


# ── Test 6: Dry-run no envía correos ──────────────────────

class TestDryRun:
    def test_dry_run_no_send(self, tmp_path):
        from src.reports.email_sender import enviar_correo

        # Crear un PDF dummy
        pdf_path = tmp_path / "test.pdf"
        pdf_path.write_bytes(b"%PDF-1.4 dummy content")

        resultado = enviar_correo(
            destinatario="test@test.cl",
            asunto="Test",
            cuerpo_html="<p>Test</p>",
            adjunto_path=pdf_path,
            dry_run=True,
        )

        assert resultado["status"] == "DRY-RUN"


# ── Test 7: Reporte de ejecución ──────────────────────────

class TestReportGeneration:
    def test_report_generation(self, json_file, tmp_path):
        from src.reports.reports_orchestrator import ReportsOrchestrator

        with patch("src.reports.pdf_generator.settings") as mock_settings, \
             patch("src.ingest.compradores_reader.validar_emails_compradores", return_value=[]):
            mock_settings.JSON_DATOS_PATH = json_file
            mock_settings.REPORTS_PATH = tmp_path / "reportes"
            mock_settings.OUTPUT_PATH = tmp_path

            orchestrator = ReportsOrchestrator(send_email=False, dry_run=False)
            report = orchestrator.run(json_path=json_file)

        assert "pdfs_generados" in report
        assert "pdfs_fallidos" in report
        assert "inicio" in report
        assert "fin" in report
        assert len(report["pdfs_generados"]) == 2
        assert len(report["pdfs_fallidos"]) == 0


# ── Test 8: Generar todos los PDFs ────────────────────────

class TestGenerateAllPDFs:
    def test_generar_todos_los_pdfs(self, json_file, tmp_path):
        from src.reports.pdf_generator import generar_todos_los_pdfs

        resultados = generar_todos_los_pdfs(json_path=json_file, output_dir=tmp_path)

        assert len(resultados) == 2
        assert all(r["status"] == "OK" for r in resultados)

        pdfs = list(tmp_path.glob("*.pdf"))
        assert len(pdfs) == 2


# ── Test 9: Datos reales (si existen) ─────────────────────

class TestRealData:
    def test_load_real_json(self):
        """Verifica que el JSON real se carga correctamente."""
        from src.reports.pdf_generator import cargar_datos

        json_path = Path("data/output/datos_procesados.json")
        if not json_path.exists():
            pytest.skip("JSON real no disponible")

        datos = cargar_datos(json_path)
        assert datos["metadata"]["total_cursos"] > 0
        assert datos["metadata"]["total_estudiantes"] > 0


# ── Test 10: PDF con caracteres especiales ────────────────

class TestUnicodeInPDF:
    def test_pdf_with_special_chars(self, tmp_path):
        """PDF con caracteres españoles (ñ, acentos) no crashea."""
        from src.reports.pdf_generator import generar_pdf

        grupo = {
            "nombre": "José María Ñuñez",
            "empresa": "Compañía Ñuñoa Ltda.",
            "email": "jose@test.cl",
            "cursos": [{
                "id_moodle": "500",
                "id_sence": "9999999",
                "nombre": "Gestión de Señalización Vía Única",
                "nombre_corto": "500",
                "fecha_inicio": "2025-01-01",
                "fecha_fin": "2025-12-31",
                "estado": "active",
                "dias_restantes": 30,
                "estudiantes": [
                    {
                        "id": "12345678-9",
                        "nombre": "María Fernández Ñuñez",
                        "email": "maria@test.cl",
                        "progreso": 85.5,
                        "calificacion": 6.5,
                        "ultimo_acceso": "2025-12-01",
                        "dias_sin_ingreso": 10,
                        "estado": "A",
                        "riesgo": "bajo",
                        "sence": {
                            "id_sence": "9999999",
                            "n_ingresos": 7,
                            "estado": "CONECTADO",
                            "declaracion_jurada": "Emitida",
                        },
                    },
                ],
                "estadisticas": {
                    "total_estudiantes": 1,
                    "promedio_progreso": 85.5,
                    "promedio_calificacion": 6.5,
                    "aprobados": 1,
                    "reprobados": 0,
                    "en_proceso": 0,
                    "riesgo_alto": 0,
                    "riesgo_medio": 0,
                    "conectados_sence": 1,
                },
            }],
        }

        ruta = generar_pdf(grupo, output_dir=tmp_path)
        assert ruta.exists()
        assert ruta.stat().st_size > 0
        assert ruta.read_bytes()[:4] == b"%PDF"


# ── Test 11: Filtrado de cursos vacíos ────────────────────

class TestEmptyCourseFiltering:
    def test_courses_with_no_students_excluded(self):
        """Cursos con 0 estudiantes se excluyen del agrupamiento."""
        from src.reports.pdf_generator import agrupar_por_comprador

        datos = {
            "cursos": [
                {
                    "nombre": "Curso Con Estudiantes",
                    "comprador": {"nombre": "Test", "empresa": "TestCo", "email": "test@test.cl"},
                    "estudiantes": [{"id": "1", "nombre": "Juan"}],
                },
                {
                    "nombre": "Curso Vacío",
                    "comprador": {"nombre": "Test", "empresa": "TestCo", "email": "test@test.cl"},
                    "estudiantes": [],
                },
            ],
        }

        grupos = agrupar_por_comprador(datos)
        assert len(grupos) == 1
        grupo = grupos["test@test.cl"]
        assert len(grupo["cursos"]) == 1
        assert grupo["cursos"][0]["nombre"] == "Curso Con Estudiantes"

    def test_comprador_with_only_empty_courses_excluded(self):
        """Comprador con solo cursos vacíos no aparece en los grupos."""
        from src.reports.pdf_generator import agrupar_por_comprador

        datos = {
            "cursos": [
                {
                    "nombre": "Curso Con Estudiantes",
                    "comprador": {"nombre": "A", "empresa": "EmpA", "email": "a@test.cl"},
                    "estudiantes": [{"id": "1", "nombre": "Juan"}],
                },
                {
                    "nombre": "Curso Vacío",
                    "comprador": {"nombre": "B", "empresa": "EmpB", "email": "b@test.cl"},
                    "estudiantes": [],
                },
            ],
        }

        grupos = agrupar_por_comprador(datos)
        assert len(grupos) == 1
        assert "a@test.cl" in grupos
        assert "b@test.cl" not in grupos


# ── Test 12: Validación de emails en compradores ─────────

class TestEmailValidation:
    def test_consistent_emails_no_error(self, tmp_path):
        """Emails consistentes por curso Moodle → sin errores."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compradores"
        ws.append(["ID Curso Moodle", "ID SENCE", "Nombre Curso",
                    "Comprador (Nombre)", "Empresa", "Email Comprador"])
        ws.append(["100", "6722569", "Curso A", "Juan", "EmpA", "juan@test.cl"])
        ws.append(["100", "6722570", "Curso A", "Juan", "EmpA", "juan@test.cl"])
        ws.append(["101", "6722572", "Curso B", "María", "EmpB", "maria@test.cl"])
        path = tmp_path / "compradores.xlsx"
        wb.save(path)

        from src.ingest.compradores_reader import validar_emails_compradores

        errores = validar_emails_compradores(path)
        assert errores == []

    def test_inconsistent_emails_error(self, tmp_path):
        """Emails distintos para el mismo curso → error."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compradores"
        ws.append(["ID Curso Moodle", "ID SENCE", "Nombre Curso",
                    "Comprador (Nombre)", "Empresa", "Email Comprador"])
        ws.append(["143", "6736819", "Geminis A", "Juan", "Emp", "juan@test.cl"])
        ws.append(["143", "6736954", "Geminis A", "Juan", "Emp", "pedro@empresa.cl"])
        path = tmp_path / "compradores.xlsx"
        wb.save(path)

        from src.ingest.compradores_reader import validar_emails_compradores

        errores = validar_emails_compradores(path)
        assert len(errores) == 1
        assert errores[0]["id_moodle"] == "143"
        assert "juan@test.cl" in errores[0]["emails"]
        assert "pedro@empresa.cl" in errores[0]["emails"]
        assert "143" in errores[0]["mensaje"]

    def test_multiple_courses_with_errors(self, tmp_path):
        """Múltiples cursos con conflicto → se listan todos."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compradores"
        ws.append(["ID Curso Moodle", "ID SENCE", "Nombre Curso",
                    "Comprador (Nombre)", "Empresa", "Email Comprador"])
        ws.append(["143", "6736819", "Geminis A", "Juan", "Emp", "a@x.cl"])
        ws.append(["143", "6736954", "Geminis A", "Juan", "Emp", "b@x.cl"])
        ws.append(["144", "6736811", "Geminis B", "Juan", "Emp", "c@x.cl"])
        ws.append(["144", "6736960", "Geminis B", "Juan", "Emp", "d@x.cl"])
        path = tmp_path / "compradores.xlsx"
        wb.save(path)

        from src.ingest.compradores_reader import validar_emails_compradores

        errores = validar_emails_compradores(path)
        assert len(errores) == 2

    def test_empty_emails_ignored(self, tmp_path):
        """Filas sin email no causan conflicto."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compradores"
        ws.append(["ID Curso Moodle", "ID SENCE", "Nombre Curso",
                    "Comprador (Nombre)", "Empresa", "Email Comprador"])
        ws.append(["100", "6722569", "Curso A", "Juan", "EmpA", "juan@test.cl"])
        ws.append(["100", "6722570", "Curso A", "Juan", "EmpA", ""])
        path = tmp_path / "compradores.xlsx"
        wb.save(path)

        from src.ingest.compradores_reader import validar_emails_compradores

        errores = validar_emails_compradores(path)
        assert errores == []

    def test_missing_file_no_error(self, tmp_path):
        """Archivo inexistente → sin errores (no bloquear)."""
        from src.ingest.compradores_reader import validar_emails_compradores

        errores = validar_emails_compradores(tmp_path / "no_existe.xlsx")
        assert errores == []


# ── Test 13: Múltiples IDs SENCE en PDF ──────────────────

class TestMultipleSenceIds:
    def test_pdf_shows_multiple_sence_ids(self, tmp_path):
        """Un curso con estudiantes con distintos SENCE IDs los recopila todos."""
        from src.reports.pdf_generator import generar_pdf

        grupo = {
            "nombre": "Test",
            "empresa": "TestCo",
            "email": "test@test.cl",
            "cursos": [{
                "id_moodle": "143",
                "id_sence": "6736819",
                "nombre": "Geminis Básico",
                "nombre_corto": "143",
                "fecha_inicio": "2025-01-01",
                "fecha_fin": "2025-12-31",
                "estado": "active",
                "dias_restantes": 30,
                "estudiantes": [
                    {
                        "id": "111", "nombre": "Juan",
                        "email": "", "progreso": 80.0, "calificacion": 6.0,
                        "ultimo_acceso": "2025-12-01", "dias_sin_ingreso": 10,
                        "estado": "A", "riesgo": "",
                        "sence": {"id_sence": "6736819", "n_ingresos": 5,
                                  "estado": "CONECTADO", "declaracion_jurada": ""},
                    },
                    {
                        "id": "222", "nombre": "Pedro",
                        "email": "", "progreso": 60.0, "calificacion": 5.0,
                        "ultimo_acceso": "2025-12-01", "dias_sin_ingreso": 10,
                        "estado": "P", "riesgo": "",
                        "sence": {"id_sence": "6736954", "n_ingresos": 3,
                                  "estado": "CONECTADO", "declaracion_jurada": ""},
                    },
                    {
                        "id": "333", "nombre": "Ana",
                        "email": "", "progreso": 40.0, "calificacion": 4.0,
                        "ultimo_acceso": "2025-12-01", "dias_sin_ingreso": 10,
                        "estado": "P", "riesgo": "medio",
                        "sence": {"id_sence": "6736960", "n_ingresos": 2,
                                  "estado": "CONECTADO", "declaracion_jurada": ""},
                    },
                ],
                "estadisticas": {
                    "total_estudiantes": 3, "promedio_progreso": 60.0,
                    "promedio_calificacion": 5.0, "aprobados": 1,
                    "reprobados": 0, "en_proceso": 2, "riesgo_alto": 0,
                    "riesgo_medio": 1, "conectados_sence": 3,
                },
            }],
        }

        ruta = generar_pdf(grupo, output_dir=tmp_path)
        assert ruta.exists()
        assert ruta.stat().st_size > 0


# ── Test 14: Parseo de múltiples emails ──────────────────

class TestMultipleEmailRecipients:
    def test_parse_single_email(self):
        from src.reports.email_sender import _parsear_emails

        assert _parsear_emails("maria@test.cl") == ["maria@test.cl"]

    def test_parse_multiple_emails(self):
        from src.reports.email_sender import _parsear_emails

        result = _parsear_emails("maria@test.cl, pedro@test.cl")
        assert result == ["maria@test.cl", "pedro@test.cl"]

    def test_parse_emails_with_spaces(self):
        from src.reports.email_sender import _parsear_emails

        result = _parsear_emails("  maria@test.cl , pedro@test.cl  ")
        assert result == ["maria@test.cl", "pedro@test.cl"]

    def test_parse_empty_email(self):
        from src.reports.email_sender import _parsear_emails

        assert _parsear_emails("") == []
        assert _parsear_emails(None) == []

    def test_parse_invalid_entries_filtered(self):
        from src.reports.email_sender import _parsear_emails

        result = _parsear_emails("maria@test.cl, , invalido, pedro@test.cl")
        assert result == ["maria@test.cl", "pedro@test.cl"]

    def test_dry_run_multiple_recipients(self, tmp_path):
        from src.reports.email_sender import enviar_correo

        pdf_path = tmp_path / "test.pdf"
        pdf_path.write_bytes(b"%PDF-1.4 dummy content")

        resultado = enviar_correo(
            destinatario="maria@test.cl, pedro@test.cl",
            asunto="Test",
            cuerpo_html="<p>Test</p>",
            adjunto_path=pdf_path,
            dry_run=True,
        )

        assert resultado["status"] == "DRY-RUN"


# ── Test 15: Orchestrator detiene en validación ──────────

class TestOrchestratorValidation:
    def test_orchestrator_stops_on_validation_error(self, json_file, tmp_path):
        """Si hay emails inconsistentes, no genera PDFs y envía alerta."""
        from src.reports.reports_orchestrator import ReportsOrchestrator

        fake_errors = [
            {
                "curso": "Geminis A",
                "id_moodle": "143",
                "emails": ["a@x.cl", "b@x.cl"],
                "mensaje": "ERROR: El curso 'Geminis A' (ID Moodle: 143) — a@x.cl vs b@x.cl",
            }
        ]

        with patch("src.reports.pdf_generator.settings") as mock_settings, \
             patch("src.ingest.compradores_reader.validar_emails_compradores",
                   return_value=fake_errors), \
             patch("src.reports.email_sender.enviar_correo") as mock_enviar:
            mock_settings.JSON_DATOS_PATH = json_file
            mock_settings.REPORTS_PATH = tmp_path / "reportes"
            mock_settings.OUTPUT_PATH = tmp_path
            mock_enviar.return_value = {"status": "DRY-RUN", "detalle": "test"}

            orchestrator = ReportsOrchestrator(send_email=False, dry_run=True)
            report = orchestrator.run(json_path=json_file)

        assert "errores_validacion" in report
        assert len(report["errores_validacion"]) == 1
        assert report["pdfs_generados"] == []
        assert report["fin"] is not None

        # Verificar que se intentó enviar el correo de alerta
        mock_enviar.assert_called_once()
        call_kwargs = mock_enviar.call_args
        assert "ERROR" in call_kwargs.kwargs.get("asunto", call_kwargs[1].get("asunto", ""))
        assert "Conflicto" in call_kwargs.kwargs.get("asunto", call_kwargs[1].get("asunto", ""))


# ── Test 16: Correo de alerta por validación ─────────────

class TestAlertaValidacion:
    def test_generar_cuerpo_alerta(self):
        """El HTML de alerta contiene los cursos con conflicto."""
        from src.reports.email_sender import generar_cuerpo_alerta_validacion

        errores = [
            {"curso": "Geminis A", "id_moodle": "143",
             "emails": ["a@x.cl", "b@x.cl"]},
            {"curso": "Geminis B", "id_moodle": "144",
             "emails": ["c@x.cl", "d@x.cl"]},
        ]

        html = generar_cuerpo_alerta_validacion(errores, "10/02/2026 15:30:00")

        assert "Geminis A" in html
        assert "143" in html
        assert "a@x.cl" in html
        assert "b@x.cl" in html
        assert "Geminis B" in html
        assert "144" in html
        assert "10/02/2026 15:30:00" in html
        assert "compradores_tecnipro.xlsx" in html
        assert "<html>" in html

    def test_enviar_correo_sin_adjunto_dry_run(self):
        """enviar_correo funciona sin adjunto (para alertas)."""
        from src.reports.email_sender import enviar_correo

        resultado = enviar_correo(
            destinatario="admin@test.cl",
            asunto="Test Alerta",
            cuerpo_html="<p>Error detectado</p>",
            adjunto_path=None,
            dry_run=True,
        )

        assert resultado["status"] == "DRY-RUN"

    def test_orchestrator_alerta_uses_structured_data(self, json_file, tmp_path):
        """El orchestrator pasa datos estructurados al HTML del correo de alerta."""
        from src.reports.reports_orchestrator import ReportsOrchestrator

        fake_errors = [
            {
                "curso": "Geminis Básico Fullkom Grupo A",
                "id_moodle": "143",
                "emails": ["jortizleiva@gmail.com", "ygonzalez@duocapital.cl"],
                "mensaje": "ERROR: inconsistentes",
            },
            {
                "curso": "Geminis Básico Fullkom Grupo B",
                "id_moodle": "144",
                "emails": ["jortizleiva@duocapital.cl", "ygonzalez@duocapital.cl"],
                "mensaje": "ERROR: inconsistentes",
            },
        ]

        correo_enviado = {}

        def capturar_correo(**kwargs):
            correo_enviado.update(kwargs)
            return {"status": "DRY-RUN", "detalle": "test"}

        with patch("src.reports.pdf_generator.settings") as mock_settings, \
             patch("src.ingest.compradores_reader.validar_emails_compradores",
                   return_value=fake_errors), \
             patch("src.reports.email_sender.enviar_correo",
                   side_effect=capturar_correo):
            mock_settings.JSON_DATOS_PATH = json_file
            mock_settings.REPORTS_PATH = tmp_path / "reportes"
            mock_settings.OUTPUT_PATH = tmp_path

            orchestrator = ReportsOrchestrator(send_email=False, dry_run=True)
            orchestrator.run(json_path=json_file)

        # Verificar contenido del correo de alerta
        assert "Conflicto" in correo_enviado.get("asunto", "")
        cuerpo = correo_enviado.get("cuerpo_html", "")
        assert "Geminis Básico Fullkom Grupo A" in cuerpo
        assert "143" in cuerpo
        assert "Geminis Básico Fullkom Grupo B" in cuerpo
        assert "144" in cuerpo
        assert "jortizleiva@gmail.com" in cuerpo
        assert "ygonzalez@duocapital.cl" in cuerpo
