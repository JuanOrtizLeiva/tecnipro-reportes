"""Rutas del servidor web del dashboard."""

import json
from datetime import datetime as _dt
import logging
import secrets
import string
import time
from collections import defaultdict
from pathlib import Path

from flask import jsonify, redirect, render_template, request, session, url_for
from flask_login import current_user, login_required, login_user, logout_user

from config import settings
from src.web.auth import check_login_rate_limit, hash_password, verify_password
from src.web import password_reset, user_manager

logger = logging.getLogger(__name__)

# Rate limiting para envío de correo: máximo 10 por minuto
_email_timestamps = defaultdict(list)
RATE_LIMIT_MAX = 10
RATE_LIMIT_WINDOW = 60  # segundos

# Caché en memoria para datos_procesados.json (evita leer disco en cada request)
_datos_cache = {"data": None, "mtime": 0.0}


def _get_datos_cached(json_path):
    """Lee JSON del disco solo si el archivo cambió (check mtime)."""
    try:
        mtime = json_path.stat().st_mtime
    except FileNotFoundError:
        return None
    if _datos_cache["data"] is not None and mtime == _datos_cache["mtime"]:
        return _datos_cache["data"]
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    _datos_cache["data"] = data
    _datos_cache["mtime"] = mtime
    return data


def _check_rate_limit(ip):
    """Retorna True si el IP excedió el límite de envíos."""
    now = time.time()
    # Limpiar timestamps viejos
    _email_timestamps[ip] = [
        t for t in _email_timestamps[ip] if now - t < RATE_LIMIT_WINDOW
    ]
    if len(_email_timestamps[ip]) >= RATE_LIMIT_MAX:
        return True
    _email_timestamps[ip].append(now)
    return False


def register_routes(app):
    """Registra todas las rutas en la app Flask."""

    # ── Login / Logout ────────────────────────────────────

    @app.route("/login", methods=["GET", "POST"])
    def login():
        """Página de login."""
        # Si ya está logueado, redirigir al dashboard
        if current_user.is_authenticated:
            return redirect(url_for("index"))

        if request.method == "GET":
            return render_template("login.html")

        # POST: verificar credenciales
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")

        # Rate limiting
        client_ip = request.remote_addr or "unknown"
        if check_login_rate_limit(client_ip):
            return render_template(
                "login.html",
                error="Demasiados intentos. Espere 15 minutos.",
                email=email,
            ), 429

        if not email or not password:
            return render_template(
                "login.html",
                error="Ingrese correo y contraseña.",
                email=email,
            )

        user = verify_password(email, password)
        if user is None:
            return render_template(
                "login.html",
                error="Credenciales inválidas.",
                email=email,
            )

        login_user(user, remember=False)  # No usar remember_me cookie
        session.permanent = True  # Sesión permanente por SESSION_LIFETIME_HOURS
        logger.info("Login exitoso: %s (%s)", user.email, user.rol)
        return redirect(url_for("index"))

    @app.route("/logout")
    def logout():
        """Destruye sesión y redirige al login."""
        if current_user.is_authenticated:
            logger.info("Logout: %s", current_user.email)

        # Cerrar sesión de Flask-Login
        logout_user()

        # Limpiar sesión de Flask
        session.clear()

        # Crear respuesta y eliminar cookies explícitamente
        response = redirect(url_for("login"))
        response.delete_cookie("session")
        response.delete_cookie("remember_token")

        # También eliminar la cookie de sesión de Flask (por si tiene otro nombre)
        response.set_cookie("session", "", expires=0)

        return response

    # ── Password Reset ────────────────────────────────────

    @app.route("/forgot-password", methods=["GET", "POST"])
    def forgot_password():
        """Solicitud de recuperación de contraseña."""
        if request.method == "GET":
            # TODO: Crear template forgot_password.html
            return render_template("forgot_password.html")

        # POST: solicitar reset
        email = request.form.get("email", "").strip()
        if not email:
            return render_template(
                "forgot_password.html",
                error="Por favor ingresa tu email"
            )

        # Generar token (siempre retornar éxito para evitar enumerar usuarios)
        token = password_reset.generar_token_reset(email)

        if token:
            # Enviar email (base_url desde request)
            base_url = request.url_root.rstrip('/')
            password_reset.enviar_email_reset(email, token, base_url)

        # Siempre mostrar mensaje de éxito (seguridad)
        return render_template(
            "forgot_password.html",
            success=True,
            message="Si el email existe, recibirás un enlace de recuperación."
        )

    @app.route("/reset-password", methods=["GET", "POST"])
    def reset_password():
        """Página para restablecer contraseña con token."""
        token = request.args.get("token", "")

        if request.method == "GET":
            # Validar token
            email = password_reset.validar_token_reset(token)
            if not email:
                return render_template(
                    "reset_password.html",
                    error="El enlace de recuperación es inválido o ha expirado."
                )

            return render_template("reset_password.html", token=token)

        # POST: cambiar contraseña
        new_password = request.form.get("password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        if not new_password or not confirm_password:
            return render_template(
                "reset_password.html",
                token=token,
                error="Por favor completa ambos campos"
            )

        if new_password != confirm_password:
            return render_template(
                "reset_password.html",
                token=token,
                error="Las contraseñas no coinciden"
            )

        if len(new_password) < 6:
            return render_template(
                "reset_password.html",
                token=token,
                error="La contraseña debe tener al menos 6 caracteres"
            )

        # Validar token nuevamente
        email = password_reset.validar_token_reset(token)
        if not email:
            return render_template(
                "reset_password.html",
                error="El enlace de recuperación es inválido o ha expirado."
            )

        # Cambiar contraseña
        try:
            user_manager.change_password(email, new_password)
            password_reset.invalidar_token_reset(token)

            logger.info("Contraseña restablecida exitosamente para %s", email)

            return render_template(
                "reset_password.html",
                success=True,
                message="Contraseña restablecida exitosamente. Ya puedes iniciar sesión."
            )
        except Exception as e:
            logger.error("Error restableciendo contraseña: %s", e)
            return render_template(
                "reset_password.html",
                token=token,
                error="Error al restablecer contraseña. Intenta nuevamente."
            )

    # ── Dashboard ─────────────────────────────────────────

    @app.route("/")
    @login_required
    def index():
        """Redirects admins to hub, serves dashboard for others."""
        if current_user.rol == "admin":
            return redirect(url_for("hub"))
        return render_template("dashboard.html")

    @app.route("/hub")
    @login_required
    def hub():
        """Admin hub menu page."""
        if current_user.rol != "admin":
            return redirect(url_for("index"))
        return render_template("hub.html")

    @app.route("/dashboard")
    @login_required
    def dashboard():
        """Dashboard de cursos (accessible directly by URL)."""
        return render_template("dashboard.html")

    # ── API: info del usuario ─────────────────────────────

    @app.route("/api/me")
    @login_required
    def api_me():
        """Retorna información del usuario actual."""
        return jsonify(current_user.to_dict())

    # ── API: datos ────────────────────────────────────────

    @app.route("/api/datos")
    @login_required
    def api_datos():
        """Retorna datos_procesados.json, filtrado por rol."""
        datos = _get_datos_cached(settings.JSON_DATOS_PATH)
        if datos is None:
            return jsonify({
                "error": "No se encontró datos_procesados.json. "
                         "Ejecute el pipeline primero: python -m src.main"
            }), 404

        # Admin (or unauthenticated when LOGIN_DISABLED) ve todo
        if not current_user.is_authenticated or current_user.rol == "admin":
            return jsonify(datos)

        # Comprador: filtrar solo sus cursos
        cursos_filtrados = [
            c for c in datos.get("cursos", [])
            if int(c.get("id_moodle", 0)) in current_user.cursos
        ]
        datos_filtrados = {
            "metadata": dict(datos.get("metadata", {})),
            "cursos": cursos_filtrados,
        }
        # Recalcular metadata
        datos_filtrados["metadata"]["total_cursos"] = len(cursos_filtrados)
        datos_filtrados["metadata"]["total_estudiantes"] = sum(
            len(c.get("estudiantes", [])) for c in cursos_filtrados
        )
        return jsonify(datos_filtrados)

    # ── API: health (pública) ─────────────────────────────

    @app.route("/api/health")
    def api_health():
        """Health check — público, no requiere autenticación."""
        datos = _get_datos_cached(settings.JSON_DATOS_PATH)
        fecha_datos = datos.get("metadata", {}).get("fecha_procesamiento") if datos else None

        return jsonify({
            "status": "ok",
            "fecha_datos": fecha_datos,
        })

    # ── API: refresh datos (todos los usuarios) ──────────────────

    @app.route("/api/refresh", methods=["POST"])
    @login_required
    def api_refresh():
        """Inicia refresh de datos Moodle en segundo plano. Todos los usuarios."""
        import os
        import re
        import subprocess
        import sys
        import threading
        import uuid

        try:
            logger.info("Refresh iniciado por %s", current_user.email)

            body = request.get_json(silent=True) or {}

            if current_user.rol == "admin":
                # Admin: puede enviar course_ids específicos en el body, o None para todos
                course_ids = body.get("course_ids") or None
                if course_ids:
                    logger.info("Refresh parcial (admin eligió): %d cursos", len(course_ids))
                else:
                    logger.info("Refresh completo de todas las categorías (admin)")
            else:
                # No-admin: siempre solo sus cursos asignados
                course_ids = current_user.cursos
                if course_ids:
                    logger.info("Refresh parcial: %d cursos del usuario", len(course_ids))

            # Generar job_id único (8 chars hex)
            job_id = uuid.uuid4().hex[:8]

            # Rutas
            project_root = Path(__file__).parent.parent.parent
            script_path = project_root / "scripts" / "run_pipeline_refresh.py"
            lock_path = project_root / "data" / "output" / "pipeline_refresh.lock"

            # Lock atómico: open('x') falla si ya existe (O_CREAT|O_EXCL)
            try:
                lock_path.parent.mkdir(parents=True, exist_ok=True)
                fd = open(str(lock_path), "x")
                fd.close()
            except FileExistsError:
                logger.warning("Refresh bloqueado por lock — ya hay un proceso activo")
                return jsonify({
                    "error": "Ya hay una actualización en proceso. Espere a que termine e intente nuevamente.",
                    "status": "busy"
                }), 409
            venv_python = Path(sys.executable)

            # Propagar entorno completo (incluye .env vars)
            child_env = os.environ.copy()
            child_env["PYTHONPATH"] = str(project_root)

            # Log file del proceso hijo
            log_path = project_root / "data" / "output" / f"pipeline_refresh_{job_id}.log"
            log_path.parent.mkdir(parents=True, exist_ok=True)
            log_file = open(str(log_path), "w", encoding="utf-8")

            # Construir comando; pasar course_ids como JSON si aplica
            cmd = [str(venv_python), str(script_path), job_id]
            if course_ids:
                cmd.append(json.dumps(course_ids))

            # Iniciar proceso completamente desvinculado del padre
            process = subprocess.Popen(
                cmd,
                stdout=log_file,
                stderr=subprocess.STDOUT,
                start_new_session=True,
                cwd=str(project_root),
                env=child_env,
            )

            logger.info(
                "Pipeline refresh background PID=%d job_id=%s iniciado",
                process.pid, job_id,
            )

            # Thread daemon para evitar procesos zombie
            def _reap_child(proc, log_fh):
                try:
                    proc.wait()
                    logger.info(
                        "Pipeline refresh PID %d finalizó (exit=%d)",
                        proc.pid, proc.returncode,
                    )
                except Exception:
                    pass
                finally:
                    try:
                        log_fh.close()
                    except Exception:
                        pass

            reaper = threading.Thread(
                target=_reap_child, args=(process, log_file), daemon=True
            )
            reaper.start()

            return jsonify({
                "status": "started",
                "job_id": job_id,
                "mensaje": "Actualización iniciada en segundo plano",
            })

        except Exception as e:
            logger.error("Error iniciando refresh en background: %s", e, exc_info=True)
            return jsonify({"error": f"Error al iniciar actualización: {str(e)}"}), 500

    @app.route("/api/refresh-status/<job_id>", methods=["GET"])
    @login_required
    def api_refresh_status(job_id):
        """Consulta el estado de un job de refresh iniciado en background."""
        import re

        # Validar job_id para evitar path traversal
        if not re.match(r'^[a-f0-9]{1,32}$', job_id):
            return jsonify({"error": "job_id inválido"}), 400

        project_root = Path(__file__).parent.parent.parent
        status_path = project_root / "data" / "output" / f"refresh_status_{job_id}.json"

        if not status_path.exists():
            # El proceso aún no escribió el archivo → todavía arrancando
            return jsonify({"status": "running", "message": "Iniciando proceso..."}), 200

        try:
            data = json.loads(status_path.read_text(encoding="utf-8"))
            return jsonify(data), 200
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    # ── API: refresh COMPLETO (Moodle + SENCE, solo superadmins) ──

    @app.route("/api/refresh-full", methods=["POST"])
    @login_required
    def api_refresh_full():
        """Inicia actualización completa en segundo plano. Solo superadmins."""
        SUPERADMINS = ["ygonzalez@duocapital.cl", "jortizleiva@duocapital.cl"]

        if current_user.email not in SUPERADMINS:
            return jsonify({"error": "No autorizado. Solo administradores principales."}), 403

        try:
            import os
            import subprocess
            import sys
            import threading
            from datetime import datetime as _dt

            logger.info("Refresh COMPLETO (background) iniciado por %s", current_user.email)

            # Ruta al script de background
            project_root = Path(__file__).parent.parent.parent
            script_path = project_root / "scripts" / "run_background_refresh.py"
            venv_python = Path(sys.executable)

            if not script_path.exists():
                raise FileNotFoundError(f"Script no encontrado: {script_path}")

            # Propagar entorno completo del proceso padre (incluye .env vars)
            child_env = os.environ.copy()
            child_env["PYTHONPATH"] = str(project_root)

            # Log con timestamp para no sobrescribir ejecuciones anteriores
            ts = _dt.now().strftime("%Y%m%d_%H%M%S")
            log_dir = project_root / "data" / "output" / "logs"
            log_dir.mkdir(parents=True, exist_ok=True)
            log_path = log_dir / f"background_refresh_{ts}.log"
            log_file = open(str(log_path), "w", encoding="utf-8")

            # Iniciar proceso en background completamente desvinculado
            process = subprocess.Popen(
                [str(venv_python), str(script_path), current_user.email],
                stdout=log_file,
                stderr=subprocess.STDOUT,
                start_new_session=True,  # Desvincular del proceso padre
                cwd=str(project_root),
                env=child_env,
            )

            logger.info(
                "Proceso en background iniciado (PID: %d, log: %s)",
                process.pid, log_path,
            )

            # Thread daemon que espera al proceso hijo para evitar zombies
            def _reap_child(proc, log_fh):
                try:
                    proc.wait()
                    if proc.returncode == 0:
                        logger.info(
                            "Proceso background PID %d finalizó exitosamente", proc.pid
                        )
                    else:
                        logger.error(
                            "Proceso background PID %d finalizó con error (exit=%d)",
                            proc.pid, proc.returncode,
                        )
                except Exception as exc:
                    logger.warning("Error en reaper de proceso background: %s", exc)
                finally:
                    try:
                        log_fh.close()
                    except Exception:
                        pass

            reaper = threading.Thread(
                target=_reap_child, args=(process, log_file), daemon=True
            )
            reaper.start()

            return jsonify({
                "status": "started",
                "mensaje": "Actualización completa iniciada en segundo plano",
                "pid": process.pid,
                "log": str(log_path.relative_to(project_root)),
                "detalle": "El proceso puede tomar 5-30 minutos. Recibirá un correo cuando finalice."
            })

        except Exception as e:
            logger.error("Error iniciando refresh en background: %s", e, exc_info=True)
            return jsonify({"error": f"Error al iniciar actualización: {str(e)}"}), 500

    # ── API: enviar correo (solo admin) ───────────────────

    @app.route("/api/enviar-correo", methods=["POST"])
    @login_required
    def api_enviar_correo():
        """Envía correo a participantes seleccionados. Solo admin."""
        # Solo admin puede enviar correos
        if current_user.is_authenticated and current_user.rol != "admin":
            return jsonify({"error": "No autorizado"}), 403

        # Rate limiting
        client_ip = request.remote_addr or "unknown"
        if _check_rate_limit(client_ip):
            return jsonify({
                "error": "Demasiados envíos. Máximo 10 por minuto."
            }), 429

        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "Body JSON requerido"}), 400

        destinatarios = data.get("destinatarios", [])
        asunto = data.get("asunto", "").strip()
        cuerpo = data.get("cuerpo", "").strip()
        cc = data.get("cc", [])

        if not destinatarios:
            return jsonify({"error": "Se requiere al menos un destinatario"}), 400
        if not asunto:
            return jsonify({"error": "Se requiere asunto"}), 400
        if not cuerpo:
            return jsonify({"error": "Se requiere cuerpo del mensaje"}), 400

        # Convertir cuerpo texto plano a HTML básico
        cuerpo_html = cuerpo.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        cuerpo_html = "<p>" + cuerpo_html.replace("\n\n", "</p><p>").replace("\n", "<br>") + "</p>"

        from src.reports.email_sender import enviar_correo

        email_str = ", ".join(destinatarios)
        cc_str = ", ".join(cc) if cc else settings.EMAIL_CC

        resultado = enviar_correo(
            destinatario=email_str,
            asunto=asunto,
            cuerpo_html=cuerpo_html,
            cc=cc_str,
            dry_run=False,
        )

        if resultado["status"] == "OK":
            logger.info("Correo enviado desde dashboard a %s", destinatarios)
            return jsonify({
                "status": "ok",
                "enviados": len(destinatarios),
            })
        else:
            logger.error("Error enviando correo desde dashboard: %s", resultado["detalle"])
            return jsonify({
                "error": resultado["detalle"],
            }), 500

    # ── API: descargar Excel ──────────────────────────────

    @app.route("/api/descargar-excel")
    @login_required
    def api_descargar_excel():
        """Genera y descarga un archivo Excel con los datos visibles."""
        from datetime import datetime
        from io import BytesIO
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Cargar datos
        json_path = settings.JSON_DATOS_PATH
        if not json_path.exists():
            return jsonify({"error": "No hay datos disponibles"}), 404

        with open(json_path, "r", encoding="utf-8") as f:
            datos = json.load(f)

        # Filtrar por rol
        cursos = datos.get("cursos", [])
        if current_user.is_authenticated and current_user.rol == "comprador":
            cursos = [c for c in cursos if int(c.get("id_moodle", 0)) in current_user.cursos]

        # Filtrar por parámetro de query (cursos visibles)
        cursos_param = request.args.get("cursos", "")
        if cursos_param:
            ids_visibles = set(cursos_param.split(","))
            cursos = [c for c in cursos if c.get("id_moodle") in ids_visibles]

        if not cursos:
            return jsonify({"error": "No hay cursos para descargar"}), 404

        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto

        # Estilos
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=13)
        subtitle_font = Font(size=11, italic=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Helper para sanear nombres de hoja (Excel no permite / \ ? * [ ] :)
        def sanitizar_nombre_hoja(nombre):
            nombre = nombre.replace('/', ' ').replace('\\', ' ')
            nombre = nombre.replace('?', '').replace('*', '')
            nombre = nombre.replace('[', '(').replace(']', ')')
            nombre = nombre.replace(':', '-')
            return nombre[:31]  # Límite de Excel

        # Crear hoja índice
        ws_index = wb.create_sheet(title="Índice", index=0)
        ws_index['A1'] = "Instituto de Capacitaciones Tecnipro"
        ws_index['A1'].font = Font(bold=True, size=16)
        ws_index['A2'] = "Reporte de Capacitación"
        ws_index['A2'].font = title_font
        ws_index['A3'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
        ws_index['A3'].font = subtitle_font

        # Headers de tabla índice
        index_headers = ["N°", "ID Curso", "Nombre del Curso", "Participantes", "Progreso Promedio", "Aprobados", "Estado"]
        for col_idx, header in enumerate(index_headers, start=1):
            cell = ws_index.cell(row=5, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Crear hojas de cursos y llenar índice
        for idx, curso in enumerate(cursos, start=1):
            # Nombre de hoja: usar nombre corto o nombre completo (sanitizado)
            nombre_base = curso.get("nombre_corto") or curso.get("nombre") or curso.get("id_moodle", "Curso")
            nombre_hoja = sanitizar_nombre_hoja(nombre_base)
            ws = wb.create_sheet(title=nombre_hoja)

            # Fila en índice (5 es header, 6 es primera fila de datos)
            row_index = idx + 5
            stats = curso.get("estadisticas", {})

            # Llenar fila de índice
            ws_index.cell(row=row_index, column=1, value=idx).border = border
            ws_index.cell(row=row_index, column=2, value=curso.get("id_moodle", "")).border = border

            # Nombre con hyperlink a la hoja del curso
            cell_nombre = ws_index.cell(row=row_index, column=3, value=curso.get("nombre", "Sin nombre"))
            cell_nombre.hyperlink = f"#{nombre_hoja}!A1"
            cell_nombre.font = Font(color="0563C1", underline="single")
            cell_nombre.border = border

            ws_index.cell(row=row_index, column=4, value=stats.get("total_estudiantes", 0)).border = border
            ws_index.cell(row=row_index, column=5, value=f"{stats.get('promedio_progreso', 0):.1f}%").border = border
            ws_index.cell(row=row_index, column=6, value=stats.get("aprobados", 0)).border = border

            estado_curso = "Activo" if curso.get("dias_restantes", 0) >= 0 else "Vencido"
            ws_index.cell(row=row_index, column=7, value=estado_curso).border = border

            # Header del curso (filas 1-3)
            ws.merge_cells('A1:J1')
            ws['A1'] = curso.get("nombre", "Sin nombre")
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')

            # Link de retorno al índice (fila 2)
            ws.merge_cells('A2:J2')
            ws['A2'] = "← Volver al Índice"
            ws['A2'].hyperlink = "#Índice!A1"
            ws['A2'].font = Font(color="0563C1", underline="single", size=10)
            ws['A2'].alignment = Alignment(horizontal='center')

            # Info del curso (fila 3)
            ws.merge_cells('A3:J3')
            info_curso = f"ID Moodle: {curso.get('id_moodle', '—')} | ID SENCE: {curso.get('id_sence', '—')} | {curso.get('fecha_inicio', '—')} a {curso.get('fecha_fin', '—')}"
            ws['A3'] = info_curso
            ws['A3'].alignment = Alignment(horizontal='center')
            ws['A3'].font = subtitle_font

            # Headers de columnas (fila 5)
            headers = [
                "Nombre", "RUT", "Correo", "Progreso (%)", "Calificación",
                "Estado", "Riesgo", "Conexiones SENCE", "DJ", "Días sin acceso"
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            # Datos de estudiantes
            estudiantes = curso.get("estudiantes", [])
            for row_idx, est in enumerate(estudiantes, start=6):
                sence = est.get("sence") or {}
                estado_texto = {"A": "Aprobado", "R": "Reprobado", "P": "En proceso"}.get(est.get("estado", ""), "—")

                valores = [
                    est.get("nombre", ""),
                    est.get("id", ""),
                    est.get("email", ""),
                    est.get("progreso", 0),
                    est.get("calificacion", 0),
                    estado_texto,
                    est.get("riesgo", "").capitalize() or "—",
                    sence.get("n_ingresos", 0) if sence else 0,
                    sence.get("declaracion_jurada", "—") if sence else "—",
                    est.get("dias_sin_ingreso", 0)
                ]

                for col_idx, valor in enumerate(valores, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                    cell.border = border
                    if col_idx in (4, 5):  # Progreso y Calificación
                        cell.alignment = Alignment(horizontal='center')

            # Fila de resumen
            stats = curso.get("estadisticas", {})
            row_resumen = len(estudiantes) + 7
            ws.merge_cells(f'A{row_resumen}:B{row_resumen}')
            ws[f'A{row_resumen}'] = "RESUMEN"
            ws[f'A{row_resumen}'].font = Font(bold=True)

            ws[f'C{row_resumen}'] = f"Total: {stats.get('total_estudiantes', 0)}"
            ws[f'D{row_resumen}'] = f"Promedio: {stats.get('promedio_progreso', 0):.1f}%"
            ws[f'E{row_resumen}'] = f"Promedio: {stats.get('promedio_calificacion', 0):.1f}"
            ws[f'F{row_resumen}'] = f"A:{stats.get('aprobados', 0)} R:{stats.get('reprobados', 0)} P:{stats.get('en_proceso', 0)}"
            ws[f'G{row_resumen}'] = f"Alto:{stats.get('riesgo_alto', 0)} Medio:{stats.get('riesgo_medio', 0)}"
            ws[f'H{row_resumen}'] = f"Conectados: {stats.get('conectados_sence', 0)}"

            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 16
            ws.column_dimensions['I'].width = 8
            ws.column_dimensions['J'].width = 15

        # Ajustar anchos de columna de la hoja índice
        ws_index.column_dimensions['A'].width = 5   # N°
        ws_index.column_dimensions['B'].width = 10  # ID Curso
        ws_index.column_dimensions['C'].width = 50  # Nombre (con hyperlink)
        ws_index.column_dimensions['D'].width = 14  # Participantes
        ws_index.column_dimensions['E'].width = 18  # Progreso Promedio
        ws_index.column_dimensions['F'].width = 12  # Aprobados
        ws_index.column_dimensions['G'].width = 12  # Estado

        # Guardar en memoria (fix para gunicorn: crear nuevo BytesIO con datos completos)
        temp_output = BytesIO()
        wb.save(temp_output)
        temp_output.seek(0)

        # Crear nuevo BytesIO con los datos completos (evita problemas de ZipFile cerrado)
        excel_data = temp_output.getvalue()
        output = BytesIO(excel_data)
        output.seek(0)

        # Nombre del archivo
        fecha_str = datetime.now().strftime("%Y%m%d")
        filename = f"Reporte_Tecnipro_{fecha_str}.xlsx"

        logger.info("Generando Excel: %s cursos para %s", len(cursos), current_user.email)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    # ── API: Coordinadores Cliente (Usuarios Compradores - solo admin) ──────────────

    def _generar_password():
        """Genera contraseña aleatoria segura de 12 caracteres."""
        alphabet = string.ascii_letters + string.digits + "!@#$%&"
        return ''.join(secrets.choice(alphabet) for _ in range(12))

    def _parsear_cursos(cursos_str):
        """Parse string de cursos separados por coma a lista de ints.

        Ejemplos:
            "190, 192, 193" → [190, 192, 193]
            "190" → [190]
            "" → []
        """
        if not cursos_str or not cursos_str.strip():
            return []

        cursos = []
        for curso in cursos_str.split(","):
            curso = curso.strip()
            if curso.isdigit():
                cursos.append(int(curso))
        return cursos

    @app.route("/api/coordinadores", methods=["GET"])
    @login_required
    def api_coordinadores_list():
        """Lista todos los coordinadores (usuarios con rol=comprador). Solo admin."""
        if current_user.rol != "admin":
            return jsonify({"error": "No autorizado"}), 403

        # Leer usuarios y filtrar solo compradores
        usuarios = user_manager._load_users()
        compradores = [
            {
                "email": u["email"],
                "nombre": u["nombre"],
                "empresa": u.get("empresa", ""),
                "cursos": u.get("cursos", []),
            }
            for u in usuarios.get("usuarios", [])
            if u.get("rol") == "comprador"
        ]

        return jsonify({"coordinadores": compradores})

    @app.route("/api/coordinadores", methods=["POST"])
    @login_required
    def api_coordinadores_create():
        """Crea un nuevo coordinador (usuario comprador) con contraseña generada. Solo admin."""
        if current_user.rol != "admin":
            return jsonify({"error": "No autorizado"}), 403

        body = request.get_json(silent=True)
        if not body:
            return jsonify({"error": "Body JSON requerido"}), 400

        # Validar campos requeridos
        nombre = body.get("nombre", "").strip()
        email = body.get("email", "").strip()
        cursos_str = body.get("cursos", "").strip()  # "190, 192, 193"
        empresa = body.get("empresa", "").strip()

        if not nombre or not email or not cursos_str:
            return jsonify({"error": "Campos requeridos: nombre, email, cursos"}), 400

        # Validar email básico
        if "@" not in email:
            return jsonify({"error": "Email inválido"}), 400

        # Parsear cursos
        cursos = _parsear_cursos(cursos_str)
        if not cursos:
            return jsonify({"error": "Debe especificar al menos un curso válido"}), 400

        # Verificar si el usuario ya existe
        if user_manager._find_user_data(email):
            return jsonify({"error": f"Ya existe un usuario con el email {email}"}), 409

        # Generar contraseña aleatoria
        password = _generar_password()

        # Crear usuario comprador
        try:
            user_manager.add_user(
                email=email,
                nombre=nombre,
                rol="comprador",
                password=password,
                cursos=cursos,
                empresa=empresa
            )
        except Exception as e:
            logger.error("Error creando coordinador: %s", e)
            return jsonify({"error": f"Error creando usuario: {str(e)}"}), 500

        # Enviar email automático con credenciales
        base_url = request.url_root.rstrip('/')
        email_sent = password_reset.enviar_email_credenciales(email, nombre, password, base_url)
        if not email_sent:
            logger.warning("No se pudo enviar email de credenciales a %s", email)

        logger.info(
            "Coordinador creado por %s: %s (%s) con cursos %s (email enviado: %s)",
            current_user.email, nombre, email, cursos, email_sent
        )

        return jsonify({
            "status": "ok",
            "coordinador": {
                "email": email,
                "nombre": nombre,
                "empresa": empresa,
                "cursos": cursos,
            },
            "password_generada": password  # Solo se muestra UNA VEZ
        }), 201

    @app.route("/api/coordinadores/<email>", methods=["PUT"])
    @login_required
    def api_coordinadores_update(email):
        """Actualiza un coordinador existente. Solo admin."""
        if current_user.rol != "admin":
            return jsonify({"error": "No autorizado"}), 403

        body = request.get_json(silent=True)
        if not body:
            return jsonify({"error": "Body JSON requerido"}), 400

        # Verificar que el usuario existe y es comprador
        user_data = user_manager._find_user_data(email)
        if not user_data:
            return jsonify({"error": "Coordinador no encontrado"}), 404

        if user_data.get("rol") != "comprador":
            return jsonify({"error": "El usuario no es un coordinador"}), 400

        # Cargar todos los usuarios
        data = user_manager._load_users()

        # Buscar y actualizar el usuario
        for u in data["usuarios"]:
            if u["email"].lower() == email.lower():
                # Actualizar campos si se proveen
                if "nombre" in body:
                    u["nombre"] = body["nombre"].strip()
                if "empresa" in body:
                    u["empresa"] = body["empresa"].strip()
                if "cursos" in body:
                    # Puede ser string "190, 192" o array [190, 192]
                    if isinstance(body["cursos"], str):
                        u["cursos"] = _parsear_cursos(body["cursos"])
                    elif isinstance(body["cursos"], list):
                        u["cursos"] = [int(c) for c in body["cursos"] if str(c).isdigit()]

                user_manager._save_users(data)

                logger.info("Coordinador %s actualizado por %s", email, current_user.email)
                return jsonify({
                    "status": "ok",
                    "coordinador": {
                        "email": u["email"],
                        "nombre": u["nombre"],
                        "empresa": u.get("empresa", ""),
                        "cursos": u.get("cursos", []),
                    }
                })

        return jsonify({"error": "Error actualizando coordinador"}), 500

    @app.route("/api/coordinadores/<email>", methods=["DELETE"])
    @login_required
    def api_coordinadores_delete(email):
        """Elimina un coordinador (usuario comprador). Solo admin."""
        if current_user.rol != "admin":
            return jsonify({"error": "No autorizado"}), 403

        # Verificar que el usuario existe y es comprador
        user_data = user_manager._find_user_data(email)
        if not user_data:
            return jsonify({"error": "Coordinador no encontrado"}), 404

        if user_data.get("rol") != "comprador":
            return jsonify({"error": "El usuario no es un coordinador"}), 400

        # Eliminar usuario
        try:
            user_manager.remove_user(email)
            logger.info("Coordinador %s eliminado por %s", email, current_user.email)
            return jsonify({"status": "ok", "mensaje": "Coordinador eliminado"})
        except Exception as e:
            logger.error("Error eliminando coordinador: %s", e)
            return jsonify({"error": f"Error eliminando usuario: {str(e)}"}), 500

    @app.route("/api/coordinadores/<email>/cursos", methods=["POST"])
    @login_required
    def api_coordinadores_add_curso(email):
        """Agrega un curso a un coordinador. Solo admin."""
        if current_user.rol != "admin":
            return jsonify({"error": "No autorizado"}), 403

        body = request.get_json(silent=True)
        if not body or "curso_id" not in body:
            return jsonify({"error": "Campo requerido: curso_id"}), 400

        curso_id = body["curso_id"]
        if not str(curso_id).isdigit():
            return jsonify({"error": "curso_id debe ser un número"}), 400

        curso_id = int(curso_id)

        # Verificar que el usuario existe y es comprador
        user_data = user_manager._find_user_data(email)
        if not user_data:
            return jsonify({"error": "Coordinador no encontrado"}), 404

        if user_data.get("rol") != "comprador":
            return jsonify({"error": "El usuario no es un coordinador"}), 400

        # Agregar curso
        try:
            user_manager.add_curso(email, curso_id)
            logger.info("Curso %d agregado a %s por %s", curso_id, email, current_user.email)

            # Recargar usuario actualizado
            user_data = user_manager._find_user_data(email)
            return jsonify({
                "status": "ok",
                "coordinador": {
                    "email": user_data["email"],
                    "nombre": user_data["nombre"],
                    "empresa": user_data.get("empresa", ""),
                    "cursos": user_data.get("cursos", []),
                }
            })
        except Exception as e:
            logger.error("Error agregando curso: %s", e)
            return jsonify({"error": f"Error agregando curso: {str(e)}"}), 500

    @app.route("/api/coordinadores/<email>/cursos/<int:curso_id>", methods=["DELETE"])
    @login_required
    def api_coordinadores_remove_curso(email, curso_id):
        """Quita un curso de un coordinador. Solo admin."""
        if current_user.rol != "admin":
            return jsonify({"error": "No autorizado"}), 403

        # Verificar que el usuario existe y es comprador
        user_data = user_manager._find_user_data(email)
        if not user_data:
            return jsonify({"error": "Coordinador no encontrado"}), 404

        if user_data.get("rol") != "comprador":
            return jsonify({"error": "El usuario no es un coordinador"}), 400

        # Cargar todos los usuarios
        data = user_manager._load_users()

        # Buscar y actualizar el usuario
        for u in data["usuarios"]:
            if u["email"].lower() == email.lower():
                cursos = u.get("cursos", [])
                if curso_id not in cursos:
                    return jsonify({"error": f"El curso {curso_id} no está asignado"}), 404

                cursos.remove(curso_id)
                u["cursos"] = cursos
                user_manager._save_users(data)

                logger.info("Curso %d quitado de %s por %s", curso_id, email, current_user.email)
                return jsonify({
                    "status": "ok",
                    "coordinador": {
                        "email": u["email"],
                        "nombre": u["nombre"],
                        "empresa": u.get("empresa", ""),
                        "cursos": u.get("cursos", []),
                    }
                })

        return jsonify({"error": "Error quitando curso"}), 500

    @app.route("/licitaciones")
    @login_required
    def licitaciones_dashboard():
        return render_template("licitaciones.html")

    @app.route("/api/licitaciones-data")
    @login_required
    def licitaciones_data():
        json_path = Path("/root/tecnipro-reportes/data/licitaciones/licitaciones_data.json")
        if not json_path.exists():
            return jsonify({"error": "Datos no disponibles aún"}), 404
        try:
            return jsonify(json.loads(json_path.read_text(encoding="utf-8")))
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    # ─── Licitacion Estado Endpoints ───



    # ─── Licitacion Notas Endpoints ───
    NOTAS_FILE = Path("/root/tecnipro-reportes/data/licitaciones/notas_oportunidades.json")

    @app.route("/api/licitacion-notas")
    @login_required
    def api_licitacion_notas():
        """Returns saved notes for all opportunities."""
        if not NOTAS_FILE.exists():
            return jsonify({})
        try:
            return jsonify(json.loads(NOTAS_FILE.read_text(encoding="utf-8")))
        except Exception:
            return jsonify({})

    @app.route("/api/licitacion-nota", methods=["POST"])
    @login_required
    def api_licitacion_nota():
        """Save or remove a note for an opportunity."""
        body = request.get_json(silent=True)
        if not body or "codigo" not in body:
            return jsonify({"error": "Datos incompletos"}), 400

        codigo = body["codigo"]
        texto = body.get("texto", "").strip()

        # Load existing notes
        notas = {}
        if NOTAS_FILE.exists():
            try:
                notas = json.loads(NOTAS_FILE.read_text(encoding="utf-8"))
            except Exception:
                notas = {}

        if texto:
            # Save or update note
            notas[codigo] = {
                "texto": texto,
                "usuario": current_user.email,
                "fecha": _dt.now().isoformat()
            }
        else:
            # Remove note
            notas.pop(codigo, None)

        # Write back
        NOTAS_FILE.parent.mkdir(parents=True, exist_ok=True)
        NOTAS_FILE.write_text(json.dumps(notas, ensure_ascii=False, indent=2), encoding="utf-8")
        return jsonify({"ok": True})

    # ── Licitacion Estado Endpoints (v2 - with history) ──
    ESTADOS_FILE = Path("/root/tecnipro-reportes/data/licitaciones/estados_oportunidades.json")

    @app.route("/api/licitacion-estados")
    @login_required
    def api_licitacion_estados():
        """Returns saved opportunity states with full history."""
        if current_user.rol != "admin":
            return jsonify({"error": "No autorizado"}), 403
        if not ESTADOS_FILE.exists():
            return jsonify({})
        try:
            import fcntl
            with open(ESTADOS_FILE, "r", encoding="utf-8") as f:
                fcntl.flock(f, fcntl.LOCK_SH)
                data = json.loads(f.read() or "{}")
                fcntl.flock(f, fcntl.LOCK_UN)
            return jsonify(data)
        except Exception:
            return jsonify({})

    @app.route("/api/licitacion-estado", methods=["POST"])
    @login_required
    def api_licitacion_estado():
        """Save a new estado entry with mandatory note. Entries are immutable."""
        if current_user.rol != "admin":
            return jsonify({"error": "No autorizado"}), 403
        body = request.get_json(silent=True)
        if not body or "codigo" not in body:
            return jsonify({"error": "Datos incompletos"}), 400

        codigo = body["codigo"]
        estado = body.get("estado", "").strip()
        nota = body.get("nota", "").strip()

        # Validation
        if not isinstance(codigo, str) or not isinstance(estado, str) or not isinstance(nota, str):
            return jsonify({"error": "Tipos invalidos"}), 400
        if not estado:
            return jsonify({"error": "Estado es obligatorio"}), 400
        if not nota:
            return jsonify({"error": "La nota es obligatoria para cambiar el estado"}), 400
        if len(codigo) > 100:
            return jsonify({"error": "Codigo demasiado largo"}), 400
        if len(nota) > 5000:
            return jsonify({"error": "Nota demasiado larga (max 5000 caracteres)"}), 400

        valid_estados = ["sin_estado", "postulando", "ganada", "perdida", "no_postulado"]
        if estado not in valid_estados:
            return jsonify({"error": "Estado no valido"}), 400

        import fcntl
        ESTADOS_FILE.parent.mkdir(parents=True, exist_ok=True)
        if not ESTADOS_FILE.exists():
            ESTADOS_FILE.write_text("{}", encoding="utf-8")

        with open(ESTADOS_FILE, "r+", encoding="utf-8") as f:
            fcntl.flock(f, fcntl.LOCK_EX)
            try:
                raw = f.read()
                estados = json.loads(raw) if raw.strip() else {}
            except Exception:
                estados = {}

            entry = {
                "estado": estado,
                "nota": nota,
                "usuario": current_user.email,
                "fecha": _dt.now().isoformat()
            }

            if codigo not in estados:
                estados[codigo] = {"estado_actual": estado, "historial": []}

            estados[codigo]["estado_actual"] = estado
            estados[codigo]["historial"].append(entry)

            f.seek(0)
            f.truncate()
            f.write(json.dumps(estados, ensure_ascii=False, indent=2))
            fcntl.flock(f, fcntl.LOCK_UN)

        return jsonify({"ok": True})

